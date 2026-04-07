from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db import IntegrityError
from django.db.models import Count, Q, Sum, Avg
from django.utils import timezone
from django.http import JsonResponse, HttpResponse
from datetime import datetime, date, timedelta
from dateutil.relativedelta import relativedelta
from .models import Member, Attendance, CulteSession, Church, GrowthConfig, AbsenceAlert, Location, UserProfile, Card, CulteType
from .forms import MemberForm, AttendanceForm, CardForm, CulteSessionForm


def user_is_admin(user):
    """Vérifie si l'utilisateur est ADMIN_CHURCH"""
    try:
        return user.profile.role == 'ADMIN_CHURCH'
    except UserProfile.DoesNotExist:
        return False


def login_view(request):
    """
    Page de connexion avec vérification de la paroisse.
    L'utilisateur doit sélectionner sa paroisse correspondant à son profil.
    """
    # Si déjà connecté, rediriger vers l'accueil
    if request.user.is_authenticated:
        return redirect('home')

    # Récupérer toutes les églises
    churches = Church.objects.select_related('location').all()

    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        church_id = request.POST.get('church')

        # Vérifier que la paroisse est sélectionnée
        if not church_id:
            messages.error(request, '⚠️ Veuillez sélectionner votre paroisse sur la carte ou dans la liste.')
            return render(request, 'registration/login.html', {'churches': churches})

        # Authentifier l'utilisateur
        user = authenticate(request, username=username, password=password)

        if user is not None:
            try:
                # Vérifier le profil utilisateur
                profile = user.profile
                selected_church = Church.objects.get(id=church_id)

                # Vérifier que la paroisse sélectionnée correspond au profil
                if profile.church == selected_church:
                    login(request, user)
                    messages.success(request, f'✅ Bienvenue {user.username} !')
                    return redirect('home')
                else:
                    messages.error(
                        request, 
                        f'❌ Votre paroisse est <b>{profile.church.name}</b> ({profile.church.location.name}). '
                        f'Veuillez sélectionner cette paroisse pour vous connecter.'
                    )
            except UserProfile.DoesNotExist:
                messages.error(request, '❌ Profil utilisateur non configuré. Contactez l\'administrateur.')
            except Church.DoesNotExist:
                messages.error(request, '❌ Paroisse invalide.')
        else:
            messages.error(request, '❌ Nom d\'utilisateur ou mot de passe incorrect.')

    return render(request, 'registration/login.html', {'churches': churches})


def logout_view(request):
    """Déconnexion de l'utilisateur"""
    logout(request)
    messages.info(request, '👋 Vous avez été déconnecté avec succès.')
    return redirect('login')


@login_required(login_url='login')
def home(request):
    """Page d'accueil avec statistiques - ADMIN seulement"""
    if not user_is_admin(request.user):
        messages.error(request, '⚠️ Accès réservé aux administrateurs uniquement.')
        return redirect('member_list')

    # Récupérer l'église de l'admin connecté
    user_church = request.user.profile.church

    context = {
        'total_members': Member.objects.filter(church=user_church).count(),
        'visitors': Member.objects.filter(church=user_church, status='VISITOR').count(),
        'regular_members': Member.objects.filter(church=user_church, status='REGULAR').count(),
        'total_cultes': CulteSession.objects.filter(church=user_church).count(),
        'attendances_today': Attendance.objects.filter(member__church=user_church, recorded_at__date=timezone.now().date()).count(),
        'user_church': user_church,
    }
    return render(request, 'church/home.html', context)


@login_required(login_url='login')
def member_list(request):
    """Liste des membres - Filtré par église de l'utilisateur"""
    # Récupérer l'église de l'utilisateur connecté
    user_church = request.user.profile.church

    # Vérifier si export EXCEL demandé (moderne et stylé)
    if request.GET.get('export') == 'csv':
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
        from openpyxl.utils import get_column_letter
        from django.http import HttpResponse

        members = Member.objects.filter(church=user_church).select_related('church')

        # Créer le workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Membres"

        # === STYLES ===
        # Header style
        header_font = Font(name='Segoe UI', bold=True, size=11, color='FFFFFF')
        header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Title style
        title_font = Font(name='Segoe UI', bold=True, size=16, color='1E40AF')
        subtitle_font = Font(name='Segoe UI', size=10, color='6B7280')
        
        # Data style
        data_font = Font(name='Segoe UI', size=10, color='1F2937')
        data_alignment = Alignment(horizontal='left', vertical='center')
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        # Alternating row fills
        light_fill = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
        white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        
        # Border
        thin_border = Border(
            left=Side(style='thin', color='E5E7EB'),
            right=Side(style='thin', color='E5E7EB'),
            top=Side(style='thin', color='E5E7EB'),
            bottom=Side(style='thin', color='E5E7EB')
        )

        # === HEADER ROWS ===
        # Row 1: Title
        ws.merge_cells('A1:G1')
        ws['A1'] = 'DHAVANT CROISSANCE'
        ws['A1'].font = Font(name='Segoe UI', bold=True, size=20, color='1E40AF')
        ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[1].height = 35
        
        # Row 2: Subtitle
        ws.merge_cells('A2:G2')
        ws['A2'] = f'Liste des Membres - {user_church.name}'
        ws['A2'].font = subtitle_font
        ws['A2'].alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[2].height = 25
        
        # Row 3: Date
        ws.merge_cells('A3:G3')
        ws['A3'] = f'Exporté le {timezone.now().strftime("%d/%m/%Y à %H:%M")}'
        ws['A3'].font = Font(name='Segoe UI', size=9, color='9CA3AF', italic=True)
        ws.row_dimensions[3].height = 20
        
        # Row 4: Empty spacer
        ws.row_dimensions[4].height = 10
        
        # Row 5: Column headers
        headers = ['Prénom', 'Nom', 'Genre', 'Téléphone', 'Statut', 'Église', 'Date adhésion']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        ws.row_dimensions[5].height = 30
        
        # === DATA ROWS ===
        for row_num, member in enumerate(members, 6):
            fill = light_fill if row_num % 2 == 0 else white_fill
            
            data = [
                member.first_name,
                member.last_name,
                member.get_genre_display(),
                member.phone or '',
                member.get_status_display(),
                member.church.name,
                member.date_joined.strftime('%d/%m/%Y')
            ]
            
            for col_num, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.font = data_font
                cell.fill = fill
                cell.border = thin_border
                
                # Center align specific columns
                if col_num in [3, 5, 7]:  # Genre, Statut, Date
                    cell.alignment = center_alignment
                else:
                    cell.alignment = data_alignment
            
            ws.row_dimensions[row_num].height = 22
        
        # === FOOTER ===
        last_row = 5 + len(list(members))
        ws.row_dimensions[last_row + 1].height = 10
        
        ws.merge_cells(f'A{last_row + 2}:G{last_row + 2}')
        ws[f'A{last_row + 2}'] = f'Total: {members.count()} membres'
        ws[f'A{last_row + 2}'].font = Font(name='Segoe UI', bold=True, size=11, color='1E40AF')
        
        ws.merge_cells(f'A{last_row + 3}:G{last_row + 3}')
        ws[f'A{last_row + 3}'] = '© Dhavant Croissance - Système de Gestion Intelligente'
        ws[f'A{last_row + 3}'].font = Font(name='Segoe UI', size=8, color='9CA3AF', italic=True)
        
        # === COLUMN WIDTHS ===
        column_widths = [20, 25, 12, 18, 15, 25, 18]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = width
        
        # === RESPONSE ===
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"Membres_{user_church.name}_{timezone.now().strftime('%Y%m%d')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response

    # Filtrer les membres de cette église uniquement avec annotation du nombre de présences
    from django.db.models import Count
    members = Member.objects.filter(
        church=user_church
    ).select_related('church').annotate(
        attendance_count=Count('history')
    )

    # Filtres
    status = request.GET.get('status')
    search = request.GET.get('search')

    if status:
        members = members.filter(status=status)
    if search:
        members = members.filter(
            first_name__icontains=search
        ) | members.filter(
            last_name__icontains=search
        )

    # Calculer les statistiques
    total = members.count()
    visitors = members.filter(status='VISITOR').count()
    regular = members.filter(status='REGULAR').count()

    context = {
        'members': members,
        'user_church': user_church,
        'total_members': total,
        'visitors_count': visitors,
        'regular_count': regular,
    }
    return render(request, 'church/member_list.html', context)


@login_required(login_url='login')
def member_create(request):
    """Créer un nouveau membre - Église pré-remplie automatiquement"""
    # Récupérer l'église de l'utilisateur connecté
    user_church = request.user.profile.church
    
    if request.method == 'POST':
        form = MemberForm(request.POST)
        if form.is_valid():
            member = form.save(commit=False)
            member.church = user_church  # Force l'église de l'utilisateur
            member.save()
            messages.success(request, f'✅ Membre {member.first_name} {member.last_name} ajouté avec succès !')
            return redirect('member_list')
    else:
        form = MemberForm()
    
    return render(request, 'church/member_form.html', {
        'form': form, 
        'title': 'Nouveau Membre',
        'user_church': user_church
    })


@login_required(login_url='login')
def member_detail(request, pk):
    """Détail d'un membre avec historique de présences"""
    member = get_object_or_404(Member, pk=pk)
    # Use 'history' related_name (defined in Attendance model)
    attendances_qs = member.history.select_related('culte_session__culte_type').order_by('-recorded_at')
    
    # Pagination
    from django.core.paginator import Paginator
    paginator = Paginator(attendances_qs, 10)  # 10 attendances per page
    page = request.GET.get('page', 1)
    attendances = paginator.get_page(page)

    context = {
        'member': member,
        'attendances': attendances,
        'page_obj': attendances,
        'is_paginated': paginator.num_pages > 1,
    }
    return render(request, 'church/member_detail.html', context)


@login_required(login_url='login')
def attendance_create(request):
    """Créer une présence pour un membre - Filtré par église"""
    user_church = request.user.profile.church
    selected_member = None
    selected_session = None
    selected_date = timezone.now().date()
    culte_sessions = []

    if request.method == 'POST':
        member_id = request.POST.get('member')
        culte_session_id = request.POST.get('culte_session')

        if member_id and culte_session_id:
            member = get_object_or_404(Member, pk=member_id, church=user_church)
            culte_session = get_object_or_404(CulteSession, pk=culte_session_id, church=user_church)

            # Vérifier si le membre est déjà présent pour cette session
            existing_attendance = Attendance.objects.filter(
                member=member,
                culte_session=culte_session
            ).first()

            if existing_attendance:
                messages.error(
                    request,
                    f'⚠️ {member.first_name} {member.last_name} est déjà présent(e) pour cette session de culte.'
                )
            else:
                attendance = Attendance.objects.create(member=member, culte_session=culte_session)
                messages.success(request, f'✅ Présence enregistrée pour {attendance.member.first_name} !')
                return redirect('member_detail', pk=attendance.member.pk)

    form = AttendanceForm()

    # Recherche de membre - limité à l'église de l'utilisateur
    search = request.GET.get('search', '')
    members = None
    if search:
        members = Member.objects.filter(church=user_church).filter(
            first_name__icontains=search
        ) | Member.objects.filter(church=user_church).filter(
            last_name__icontains=search
        )
        # Add pagination
        from django.core.paginator import Paginator
        paginator = Paginator(members.order_by('last_name', 'first_name'), 10)  # 10 members per page
        page = request.GET.get('page', 1)
        members = paginator.get_page(page)
    else:
        members = Member.objects.filter(church=user_church)[:10]

    # Get selected member info
    if request.GET.get('member_id'):
        selected_member = get_object_or_404(Member, pk=request.GET.get('member_id'))

    # Get date for filtering
    if request.GET.get('date'):
        selected_date = datetime.strptime(request.GET.get('date'), '%Y-%m-%d').date()

    # Get culte sessions for selected date and church
    if selected_date:
        culte_sessions = CulteSession.objects.filter(
            church=user_church,
            date__date=selected_date
        ).select_related('culte_type')

    context = {
        'form': form,
        'search': search,
        'members': members,
        'user_church': user_church,
        'selected_member': selected_member,
        'selected_date': selected_date,
        'culte_sessions': culte_sessions,
        'is_paginated': hasattr(members, 'has_next'),
        'page_obj': members if hasattr(members, 'has_next') else None,
    }
    return render(request, 'church/attendance_form.html', context)


@login_required(login_url='login')
def attendance_export(request):
    """Exporter l'historique des présences en Excel stylé"""
    user_church = request.user.profile.church

    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse

    attendances = Attendance.objects.filter(
        member__church=user_church
    ).select_related('member', 'culte_session__culte_type').order_by('-culte_session__date')

    # Créer le workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Présences"

    # === STYLES ===
    header_font = Font(name='Segoe UI', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='10B981', end_color='10B981', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    title_font = Font(name='Segoe UI', bold=True, size=20, color='065F46')
    subtitle_font = Font(name='Segoe UI', size=10, color='6B7280')
    
    data_font = Font(name='Segoe UI', size=10, color='1F2937')
    data_alignment = Alignment(horizontal='left', vertical='center')
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    light_fill = PatternFill(start_color='F0FDF4', end_color='F0FDF4', fill_type='solid')
    white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    
    thin_border = Border(
        left=Side(style='thin', color='E5E7EB'),
        right=Side(style='thin', color='E5E7EB'),
        top=Side(style='thin', color='E5E7EB'),
        bottom=Side(style='thin', color='E5E7EB')
    )

    # === HEADER ===
    ws.merge_cells('A1:E1')
    ws['A1'] = 'DHAVANT CROISSANCE'
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 35
    
    ws.merge_cells('A2:E2')
    ws['A2'] = f'Historique des Présences - {user_church.name}'
    ws['A2'].font = subtitle_font
    ws['A2'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[2].height = 25
    
    ws.merge_cells('A3:E3')
    ws['A3'] = f'Exporté le {timezone.now().strftime("%d/%m/%Y à %H:%M")}'
    ws['A3'].font = Font(name='Segoe UI', size=9, color='9CA3AF', italic=True)
    ws.row_dimensions[3].height = 20
    ws.row_dimensions[4].height = 10
    
    # Column headers
    headers = ['Date', 'Membre', 'Type de Culte', 'Thème', 'Enregistré le']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    ws.row_dimensions[5].height = 30
    
    # === DATA ===
    for row_num, attendance in enumerate(attendances, 6):
        fill = light_fill if row_num % 2 == 0 else white_fill
        
        data = [
            attendance.culte_session.date.strftime('%d/%m/%Y %H:%M'),
            f'{attendance.member.first_name} {attendance.member.last_name}',
            attendance.culte_session.culte_type.name,
            attendance.culte_session.theme or '',
            attendance.recorded_at.strftime('%d/%m/%Y %H:%M')
        ]
        
        for col_num, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.font = data_font
            cell.fill = fill
            cell.border = thin_border
            
            if col_num in [1, 3, 5]:  # Date, Type, Enregistré le
                cell.alignment = center_alignment
            else:
                cell.alignment = data_alignment
        
        ws.row_dimensions[row_num].height = 22
    
    # === FOOTER ===
    last_row = 5 + len(list(attendances))
    ws.row_dimensions[last_row + 1].height = 10
    
    ws.merge_cells(f'A{last_row + 2}:E{last_row + 2}')
    ws[f'A{last_row + 2}'] = f'Total: {attendances.count()} présences enregistrées'
    ws[f'A{last_row + 2}'].font = Font(name='Segoe UI', bold=True, size=11, color='065F46')
    
    ws.merge_cells(f'A{last_row + 3}:E{last_row + 3}')
    ws[f'A{last_row + 3}'] = '© Dhavant Croissance - Système de Gestion Intelligente'
    ws[f'A{last_row + 3}'].font = Font(name='Segoe UI', size=8, color='9CA3AF', italic=True)
    
    # === COLUMN WIDTHS ===
    column_widths = [22, 30, 20, 35, 22]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width
    
    # === RESPONSE ===
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"Presences_{user_church.name}_{timezone.now().strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


@login_required(login_url='login')
def alerts_export(request):
    """Exporter les alertes d'absence en Excel stylé"""
    user_church = request.user.profile.church

    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse

    alerts = AbsenceAlert.objects.filter(
        church=user_church
    ).select_related('member', 'church').order_by('-created_at')

    # Créer le workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Alertes Absences"

    # === STYLES ===
    header_font = Font(name='Segoe UI', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='EF4444', end_color='EF4444', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    title_font = Font(name='Segoe UI', bold=True, size=20, color='991B1B')
    subtitle_font = Font(name='Segoe UI', size=10, color='6B7280')
    
    data_font = Font(name='Segoe UI', size=10, color='1F2937')
    data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    light_fill = PatternFill(start_color='FEF2F2', end_color='FEF2F2', fill_type='solid')
    white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    
    # Status colors
    resolved_fill = PatternFill(start_color='DCFCE7', end_color='DCFCE7', fill_type='solid')
    pending_fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
    
    thin_border = Border(
        left=Side(style='thin', color='E5E7EB'),
        right=Side(style='thin', color='E5E7EB'),
        top=Side(style='thin', color='E5E7EB'),
        bottom=Side(style='thin', color='E5E7EB')
    )

    # === HEADER ===
    ws.merge_cells('A1:F1')
    ws['A1'] = 'DHAVANT CROISSANCE'
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 35
    
    ws.merge_cells('A2:F2')
    ws['A2'] = f'Alertes d\'Absences - {user_church.name}'
    ws['A2'].font = subtitle_font
    ws['A2'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[2].height = 25
    
    ws.merge_cells('A3:F3')
    ws['A3'] = f'Exporté le {timezone.now().strftime("%d/%m/%Y à %H:%M")}'
    ws['A3'].font = Font(name='Segoe UI', size=9, color='9CA3AF', italic=True)
    ws.row_dimensions[3].height = 20
    ws.row_dimensions[4].height = 10
    
    # Column headers
    headers = ['Membre', 'Église', 'Dernière présence', 'Créée le', 'Statut', 'Notes pastorales']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    ws.row_dimensions[5].height = 30
    
    # === DATA ===
    for row_num, alert in enumerate(alerts, 6):
        fill = light_fill if row_num % 2 == 0 else white_fill
        
        status_text = 'Résolu' if alert.is_resolved else 'En attente'
        status_fill = resolved_fill if alert.is_resolved else pending_fill
        
        data = [
            f'{alert.member.first_name} {alert.member.last_name}',
            alert.church.name,
            alert.last_seen.strftime('%d/%m/%Y') if alert.last_seen else '',
            alert.created_at.strftime('%d/%m/%Y %H:%M'),
            status_text,
            alert.pastoral_notes or ''
        ]
        
        for col_num, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.font = data_font
            cell.fill = fill
            cell.border = thin_border
            
            # Special styling for status column
            if col_num == 5:  # Statut
                cell.fill = status_fill
                cell.font = Font(name='Segoe UI', size=10, bold=True, 
                               color='166534' if alert.is_resolved else '991B1B')
                cell.alignment = center_alignment
            elif col_num in [3, 4]:  # Dates
                cell.alignment = center_alignment
            else:
                cell.alignment = data_alignment
        
        ws.row_dimensions[row_num].height = 25
    
    # === FOOTER ===
    last_row = 5 + len(list(alerts))
    ws.row_dimensions[last_row + 1].height = 10
    
    resolved_count = alerts.filter(is_resolved=True).count()
    pending_count = alerts.filter(is_resolved=False).count()
    
    ws.merge_cells(f'A{last_row + 2}:F{last_row + 2}')
    ws[f'A{last_row + 2}'] = f'Total: {alerts.count()} alertes ({resolved_count} résolues, {pending_count} en attente)'
    ws[f'A{last_row + 2}'].font = Font(name='Segoe UI', bold=True, size=11, color='991B1B')
    
    ws.merge_cells(f'A{last_row + 3}:F{last_row + 3}')
    ws[f'A{last_row + 3}'] = '© Dhavant Croissance - Système de Gestion Intelligente'
    ws[f'A{last_row + 3}'].font = Font(name='Segoe UI', size=8, color='9CA3AF', italic=True)
    
    # === COLUMN WIDTHS ===
    column_widths = [30, 25, 20, 22, 15, 40]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width
    
    # === RESPONSE ===
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"Alertes_{user_church.name}_{timezone.now().strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


@login_required(login_url='login')
def api_culte_sessions(request):
    """API endpoint to get culte sessions for a specific date - Filtré par église du user"""
    date_str = request.GET.get('date')

    if not date_str:
        return JsonResponse({'error': 'Date required'}, status=400)

    try:
        selected_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        user_church = request.user.profile.church

        sessions = CulteSession.objects.filter(
            church=user_church,
            date__date=selected_date
        ).select_related('culte_type')

        # Format for display
        formatted_sessions = []
        for s in sessions:
            formatted_sessions.append({
                'id': s.id,
                'theme': s.theme or 'Sans thème',
                'officiant': s.officiant or 'Sans officiant',
                'culte_type': s.culte_type.name if s.culte_type else 'Culte'
            })

        return JsonResponse({'sessions': formatted_sessions})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required(login_url='login')
def alerts_list(request):
    """Liste des alertes d'absence - ADMIN seulement, filtré par église"""
    if not user_is_admin(request.user):
        messages.error(request, '⚠️ Accès réservé aux administrateurs uniquement.')
        return redirect('member_list')

    # Récupérer l'église de l'admin connecté
    user_church = request.user.profile.church
    
    # Générer les alertes d'absence automatiquement
    generate_absence_alerts(user_church)

    alerts = AbsenceAlert.objects.select_related('member', 'church').filter(
        church=user_church,
        is_resolved=False
    ).order_by('-created_at')
    
    # Récupérer la configuration pour afficher le seuil d'absence
    config = GrowthConfig.get_config()

    context = {
        'alerts': alerts,
        'user_church': user_church,
        'config': config
    }
    return render(request, 'church/alerts_list.html', context)


def generate_absence_alerts(user_church):
    """
    Génère automatiquement les alertes d'absence pour les membres actifs
    qui n'ont pas été vus depuis le seuil défini dans GrowthConfig.
    """
    config = GrowthConfig.get_config()
    absence_threshold = timezone.now() - timedelta(days=config.absence_limit_days)
    
    # Récupérer tous les membres actifs de l'église
    active_members = Member.objects.filter(
        church=user_church,
        status='REGULAR'
    ).select_related('church')
    
    for member in active_members:
        # Trouver la dernière présence du membre
        last_attendance = Attendance.objects.filter(
            member=member
        ).select_related('culte_session').order_by('-culte_session__date').first()
        
        if last_attendance is None:
            # Membre sans aucune présence - ne pas créer d'alerte
            continue
        
        # Vérifier si la dernière présence est plus ancienne que le seuil
        if last_attendance.culte_session.date < absence_threshold:
            # Vérifier si une alerte existe déjà pour ce membre
            existing_alert = AbsenceAlert.objects.filter(
                member=member,
                church=user_church,
                is_resolved=False
            ).first()
            
            if not existing_alert:
                # Créer une nouvelle alerte
                AbsenceAlert.objects.create(
                    church=user_church,
                    member=member,
                    last_seen=last_attendance.recorded_at
                )


@login_required(login_url='login')
def alert_resolve(request, alert_id):
    """Marquer une alerte d'absence comme résolue"""
    if not user_is_admin(request.user):
        messages.error(request, '⚠️ Accès réservé aux administrateurs uniquement.')
        return redirect('member_list')
    
    alert = get_object_or_404(AbsenceAlert, pk=alert_id, church=request.user.profile.church)
    
    if request.method == 'POST':
        pastoral_notes = request.POST.get('pastoral_notes', '')
        alert.is_resolved = True
        alert.pastoral_notes = pastoral_notes
        alert.save()
        messages.success(request, f'✅ Alerte marquée comme résolue pour {alert.member.first_name} {alert.member.last_name}')
    
    return redirect('alerts_list')


@login_required(login_url='login')
def card_create(request, member_pk=None):
    """Créer une carte pour un membre - Filtré par église"""
    user_church = request.user.profile.church
    
    # Get member if pk provided
    member = None
    if member_pk:
        member = get_object_or_404(Member, pk=member_pk, church=user_church)
    
    if request.method == 'POST':
        form = CardForm(request.POST)
        if form.is_valid():
            card = form.save(commit=False)
            # Ensure member belongs to user's church
            if card.member.church != user_church:
                messages.error(request, '❌ Vous ne pouvez créer des cartes que pour les membres de votre église.')
                return redirect('member_list')
            card.save()
            messages.success(request, f'✅ Carte {card.card_number} créée pour {card.member.first_name} {card.member.last_name} !')
            return redirect('member_detail', pk=card.member.pk)
    else:
        form = CardForm()
        if member:
            form.fields['member'].initial = member

    # Get member's existing cards if viewing for specific member
    existing_cards = []
    if member:
        existing_cards = member.cards.all().order_by('-created_at')

    context = {
        'form': form,
        'member': member,
        'existing_cards': existing_cards,
        'user_church': user_church,
    }
    return render(request, 'church/card_form.html', context)


@login_required(login_url='login')
def culte_session_create(request):
    """Créer une session de culte - Filtré par église"""
    user_church = request.user.profile.church
    
    if request.method == 'POST':
        form = CulteSessionForm(request.POST)
        if form.is_valid():
            session = form.save(commit=False)
            session.church = user_church  # Force l'église de l'utilisateur
            try:
                session.save()
                messages.success(request, f'✅ Session de culte "{session.theme}" créée pour le {session.date.strftime("%d %b %Y")} !')
                return redirect('home')
            except IntegrityError:
                messages.error(request, '⚠️ Une session de culte existe déjà pour ce type, cette date et votre église.')
    else:
        form = CulteSessionForm()

    context = {
        'form': form,
        'user_church': user_church,
    }
    return render(request, 'church/culte_session_form.html', context)


@login_required(login_url='login')
def culte_session_list(request):
    """Liste des sessions de culte avec statistiques - Filtré par église"""
    user_church = request.user.profile.church

    # Vérifier si export EXCEL demandé (moderne et stylé)
    if request.GET.get('export') == 'csv':
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter
        from django.http import HttpResponse

        sessions = CulteSession.objects.filter(
            church=user_church
        ).select_related('culte_type').order_by('-date')

        # Créer le workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Sessions de Culte"

        # === STYLES ===
        header_font = Font(name='Segoe UI', bold=True, size=11, color='FFFFFF')
        header_fill = PatternFill(start_color='8B5CF6', end_color='8B5CF6', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        title_font = Font(name='Segoe UI', bold=True, size=20, color='5B21B6')
        subtitle_font = Font(name='Segoe UI', size=10, color='6B7280')
        
        data_font = Font(name='Segoe UI', size=10, color='1F2937')
        data_alignment = Alignment(horizontal='left', vertical='center')
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        light_fill = PatternFill(start_color='F5F3FF', end_color='F5F3FF', fill_type='solid')
        white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        
        thin_border = Border(
            left=Side(style='thin', color='E5E7EB'),
            right=Side(style='thin', color='E5E7EB'),
            top=Side(style='thin', color='E5E7EB'),
            bottom=Side(style='thin', color='E5E7EB')
        )

        # === HEADER ===
        ws.merge_cells('A1:E1')
        ws['A1'] = 'DHAVANT CROISSANCE'
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[1].height = 35
        
        ws.merge_cells('A2:E2')
        ws['A2'] = f'Sessions de Culte - {user_church.name}'
        ws['A2'].font = subtitle_font
        ws['A2'].alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[2].height = 25
        
        ws.merge_cells('A3:E3')
        ws['A3'] = f'Exporté le {timezone.now().strftime("%d/%m/%Y à %H:%M")}'
        ws['A3'].font = Font(name='Segoe UI', size=9, color='9CA3AF', italic=True)
        ws.row_dimensions[3].height = 20
        ws.row_dimensions[4].height = 10
        
        # Column headers
        headers = ['Date', 'Type', 'Thème', 'Officiant', 'Église']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        ws.row_dimensions[5].height = 30
        
        # === DATA ===
        for row_num, session in enumerate(sessions, 6):
            fill = light_fill if row_num % 2 == 0 else white_fill
            
            data = [
                session.date.strftime('%d/%m/%Y %H:%M'),
                session.culte_type.name,
                session.theme or '',
                session.officiant or '',
                session.church.name
            ]
            
            for col_num, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.font = data_font
                cell.fill = fill
                cell.border = thin_border
                
                if col_num in [1, 2]:  # Date, Type
                    cell.alignment = center_alignment
                else:
                    cell.alignment = data_alignment
            
            ws.row_dimensions[row_num].height = 22
        
        # === FOOTER ===
        last_row = 5 + len(list(sessions))
        ws.row_dimensions[last_row + 1].height = 10
        
        ws.merge_cells(f'A{last_row + 2}:E{last_row + 2}')
        ws[f'A{last_row + 2}'] = f'Total: {sessions.count()} sessions de culte'
        ws[f'A{last_row + 2}'].font = Font(name='Segoe UI', bold=True, size=11, color='5B21B6')
        
        ws.merge_cells(f'A{last_row + 3}:E{last_row + 3}')
        ws[f'A{last_row + 3}'] = '© Dhavant Croissance - Système de Gestion Intelligente'
        ws[f'A{last_row + 3}'].font = Font(name='Segoe UI', size=8, color='9CA3AF', italic=True)
        
        # === COLUMN WIDTHS ===
        column_widths = [22, 20, 40, 30, 25]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = width
        
        # === RESPONSE ===
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"Cultes_{user_church.name}_{timezone.now().strftime('%Y%m%d')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response

    # Get all culte sessions for user's church
    sessions = CulteSession.objects.filter(church=user_church).select_related('culte_type').order_by('-date')

    # Filters
    culte_type = request.GET.get('culte_type')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')

    if culte_type:
        sessions = sessions.filter(culte_type_id=culte_type)
    if date_from:
        sessions = sessions.filter(date__date__gte=date_from)
    if date_to:
        sessions = sessions.filter(date__date__lte=date_to)

    # Get culte types for filter dropdown
    culte_types = CulteType.objects.all()

    # Calculate statistics for each session
    sessions_with_stats = []
    total_presences = 0
    total_hommes = 0
    total_femmes = 0
    total_enfants = 0
    
    for session in sessions:
        stats = session.get_attendance_stats()

        # Count active members present
        active_present = Attendance.objects.filter(
            culte_session=session,
            member__status='REGULAR'
        ).count()

        # Count total active members in this church
        total_active = Member.objects.filter(church=user_church, status='REGULAR').count()

        # Calculate absent active members
        active_absent = total_active - active_present

        sessions_with_stats.append({
            'session': session,
            'total': stats['total'],
            'hommes': stats['hommes'],
            'femmes': stats['femmes'],
            'enfants': stats['enfants'],
            'actifs_presents': active_present,
            'actifs_absents': active_absent,
        })
        
        # Accumulate totals
        total_presences += stats['total'] or 0
        total_hommes += stats['hommes'] or 0
        total_femmes += stats['femmes'] or 0
        total_enfants += stats['enfants'] or 0

    context = {
        'sessions': sessions_with_stats,
        'culte_types': culte_types,
        'user_church': user_church,
        'selected_culte_type': culte_type,
        'selected_date_from': date_from,
        'selected_date_to': date_to,
        'total_presences': total_presences,
        'total_hommes': total_hommes,
        'total_femmes': total_femmes,
        'total_enfants': total_enfants,
    }
    return render(request, 'church/culte_session_list.html', context)


@login_required(login_url='login')
def statistics(request):
    """Page de statistiques avancées - Réservée aux administrateurs"""
    if not user_is_admin(request.user):
        messages.error(request, '⚠️ Accès réservé aux administrateurs uniquement.')
        return redirect('member_list')

    user_church = request.user.profile.church

    # Filtres
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    culte_type_id = request.GET.get('culte_type', '')

    # Période par défaut (30 derniers jours)
    if not date_from:
        date_from = (timezone.now().date() - timedelta(days=30)).isoformat()
    if not date_to:
        date_to = timezone.now().date().isoformat()

    # Base queryset filtrée
    sessions_qs = CulteSession.objects.filter(church=user_church)
    attendances_qs = Attendance.objects.filter(member__church=user_church)

    # Appliquer les filtres
    if date_from:
        sessions_qs = sessions_qs.filter(date__date__gte=date_from)
        attendances_qs = attendances_qs.filter(culte_session__date__date__gte=date_from)
    if date_to:
        sessions_qs = sessions_qs.filter(date__date__lte=date_to)
        attendances_qs = attendances_qs.filter(culte_session__date__date__lte=date_to)
    if culte_type_id:
        sessions_qs = sessions_qs.filter(culte_type_id=culte_type_id)
        attendances_qs = attendances_qs.filter(culte_session__culte_type_id=culte_type_id)

    # Statistiques générales
    total_members = Member.objects.filter(church=user_church).count()
    visitors = Member.objects.filter(church=user_church, status='VISITOR').count()
    regular_members = Member.objects.filter(church=user_church, status='REGULAR').count()
    total_cultes = sessions_qs.count()
    total_presences = attendances_qs.count()

    # Statistiques par genre
    hommes = Member.objects.filter(church=user_church, genre='H').count()
    femmes = Member.objects.filter(church=user_church, genre='F').count()
    enfants = Member.objects.filter(church=user_church, genre='E').count()

    # Statistiques par statut
    status_stats = {
        'VISITOR': visitors,
        'REGULAR': regular_members,
    }

    # Présences par culte type (pour diagramme)
    culte_types = CulteType.objects.all()
    presences_by_type = []
    for ct in culte_types:
        count = attendances_qs.filter(culte_session__culte_type=ct).count()
        presences_by_type.append({
            'name': ct.name,
            'count': count,
            'percentage': round((count / total_presences * 100) if total_presences > 0 else 0, 1)
        })

    # Présences par mois (pour diagramme)
    presences_by_month = []
    for i in range(11, -1, -1):
        month_date = timezone.now().date() - timedelta(days=i*30)
        month_name = month_date.strftime('%b')
        month_start = month_date.replace(day=1)
        if month_date.month == 12:
            month_end = month_date.replace(year=month_date.year+1, month=1, day=1) - timedelta(days=1)
        else:
            month_end = month_date.replace(month=month_date.month+1, day=1) - timedelta(days=1)
        
        count = attendances_qs.filter(
            culte_session__date__date__gte=month_start,
            culte_session__date__date__lte=month_end
        ).count()
        presences_by_month.append({
            'month': month_name,
            'count': count
        })

    # Présences par semaine (pour diagramme - 4 dernières semaines)
    presences_by_week = []
    for i in range(3, -1, -1):
        week_end = timezone.now().date() - timedelta(days=i*7)
        week_start = week_end - timedelta(days=6)
        count = attendances_qs.filter(
            culte_session__date__date__gte=week_start,
            culte_session__date__date__lte=week_end
        ).count()
        presences_by_week.append({
            'week': f'S{i+1}',
            'count': count
        })

    # Taux de participation - Formule MDEVISP officielle
    # Selon cahier des charges : % = (X / Y) × 100
    # où X = N + M + T (nombre total de cultes)
    # et Y = Effectif total (membres actifs)
    # INTERPRÉTATION PRATIQUE : On utilise plutôt le ratio présences/membres
    # car X (nombre de cultes) seul ne donne pas une information pertinente
    # Formule utilisée : (Total présences / (Effectif × Nb cultes)) × 100
    # Cette formule donne le pourcentage de participation MOYEN par culte
    if regular_members > 0 and total_cultes > 0:
        participation_rate = round((total_presences / (regular_members * total_cultes)) * 100, 1)
    else:
        participation_rate = 0

    context = {
        'total_members': total_members,
        'visitors': visitors,
        'regular_members': regular_members,
        'total_cultes': total_cultes,
        'total_presences': total_presences,
        'hommes': hommes,
        'femmes': femmes,
        'status_stats': status_stats,
        'presences_by_type': presences_by_type,
        'presences_by_month': presences_by_month,
        'presences_by_week': presences_by_week,
        'participation_rate': participation_rate,
        'culte_types': culte_types,
        'selected_culte_type': culte_type_id,
        'selected_date_from': date_from,
        'selected_date_to': date_to,
        'user_church': user_church,
    }

    return render(request, 'church/statistics.html', context)


@login_required(login_url='login')
def statistics_export(request):
    """Export des statistiques en Excel professionnel avec formules"""
    if not user_is_admin(request.user):
        messages.error(request, '⚠️ Accès réservé aux administrateurs uniquement.')
        return redirect('member_list')

    export_format = request.GET.get('format', 'excel')
    user_church = request.user.profile.church

    # Récupérer les mêmes données que la vue statistics
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')

    if not date_from:
        date_from = (timezone.now().date() - timedelta(days=30)).isoformat()
    if not date_to:
        date_to = timezone.now().date().isoformat()

    total_members = Member.objects.filter(church=user_church).count()
    visitors = Member.objects.filter(church=user_church, status='VISITOR').count()
    regular_members = Member.objects.filter(church=user_church, status='REGULAR').count()
    hommes = Member.objects.filter(church=user_church, genre='H').count()
    femmes = Member.objects.filter(church=user_church, genre='F').count()
    enfants = Member.objects.filter(church=user_church, genre='E').count()

    sessions_qs = CulteSession.objects.filter(church=user_church, date__date__gte=date_from, date__date__lte=date_to)
    attendances_qs = Attendance.objects.filter(member__church=user_church)
    if date_from:
        attendances_qs = attendances_qs.filter(culte_session__date__date__gte=date_from)
    if date_to:
        attendances_qs = attendances_qs.filter(culte_session__date__date__lte=date_to)

    total_cultes = sessions_qs.count()
    total_presences = attendances_qs.count()
    
    # Taux de participation
    if regular_members > 0 and total_cultes > 0:
        participation_rate = round((total_presences / (regular_members * total_cultes)) * 100, 1)
    else:
        participation_rate = 0

    # Pourcentages par genre
    pct_hommes = round((hommes / total_members * 100) if total_members > 0 else 0, 1)
    pct_femmes = round((femmes / total_members * 100) if total_members > 0 else 0, 1)
    pct_enfants = round((enfants / total_members * 100) if total_members > 0 else 0, 1)

    if export_format == 'excel':
        # Import openpyxl pour vrai Excel
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, Color
            from openpyxl.utils import get_column_letter
            from openpyxl.chart import BarChart, PieChart, LineChart, Reference
            from openpyxl.chart.label import DataLabelList
            from io import BytesIO
        except ImportError:
            messages.error(request, '❌ La librairie openpyxl n\'est pas installée. Veuillez l\'installer.')
            return redirect('statistics')

        # Création du classeur
        wb = Workbook()
        ws = wb.active
        ws.title = "Statistiques"

        # Styles
        title_font = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
        header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
        normal_font = Font(name='Calibri', size=11)
        bold_font = Font(name='Calibri', size=11, bold=True)
        
        title_fill = PatternFill(start_color='1E40AF', end_color='1E40AF', fill_type='solid')
        header_fill = PatternFill(start_color='3B82F6', end_color='3B82F6', fill_type='solid')
        light_fill = PatternFill(start_color='EFF6FF', end_color='EFF6FF', fill_type='solid')
        highlight_fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
        
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # === EN-TÊTE ===
        ws.merge_cells('A1:D1')
        cell = ws['A1']
        cell.value = f"STATISTIQUES - {user_church.name.upper()}"
        cell.font = title_font
        cell.fill = title_fill
        cell.alignment = center_align
        
        ws.merge_cells('A2:D2')
        ws['A2'].value = f"Période : du {date_from} au {date_to}"
        ws['A2'].font = Font(name='Calibri', size=11, italic=True)
        ws['A2'].alignment = center_align

        # === SECTION 1: MEMBRES ===
        ws.merge_cells('A4:D4')
        ws['A4'].value = "1. STATISTIQUES MEMBRES"
        ws['A4'].font = header_font
        ws['A4'].fill = header_fill
        ws['A4'].alignment = center_align

        headers_membres = ['Indicateur', 'Effectif', 'Pourcentage', 'Formule']
        for col, header in enumerate(headers_membres, 1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.font = bold_font
            cell.fill = PatternFill(start_color='93C5FD', end_color='93C5FD', fill_type='solid')
            cell.alignment = center_align
            cell.border = thin_border

        # Données membres
        membres_data = [
            ['Total Membres', total_members, '100%', '=B6/SUM($B$6:$B$10)*100'],
            ['Membres Actifs', regular_members, str(round((regular_members/total_members*100) if total_members > 0 else 0, 1)) + '%', '=B7/SUM($B$6:$B$10)*100'],
            ['Visiteurs', visitors, str(round((visitors/total_members*100) if total_members > 0 else 0, 1)) + '%', '=B8/SUM($B$6:$B$10)*100'],
            ['Hommes', hommes, str(pct_hommes) + '%', '=B9/$B$6*100'],
            ['Femmes', femmes, str(pct_femmes) + '%', '=B10/$B$6*100'],
            ['Enfants', enfants, str(pct_enfants) + '%', '=B11/$B$6*100'],
        ]

        for row_idx, data in enumerate(membres_data, 6):
            for col_idx, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = bold_font if col_idx == 1 else normal_font
                cell.alignment = center_align if col_idx > 1 else left_align
                cell.border = thin_border
                if col_idx == 1:
                    cell.fill = light_fill

        # === SECTION 2: CULTES ===
        ws.merge_cells('A12:D12')
        ws['A12'].value = "2. STATISTIQUES CULTES & PRÉSENCES"
        ws['A12'].font = header_font
        ws['A12'].fill = header_fill
        ws['A12'].alignment = center_align

        headers_cultes = ['Indicateur', 'Valeur', 'Formule Excel', '']
        for col, header in enumerate(headers_cultes, 1):
            cell = ws.cell(row=13, column=col, value=header)
            cell.font = bold_font
            cell.fill = PatternFill(start_color='93C5FD', end_color='93C5FD', fill_type='solid')
            cell.alignment = center_align
            cell.border = thin_border

        cultes_data = [
            ['Total Cultes', total_cultes, '=COUNTA(A14:A100)', ''],
            ['Total Présences', total_presences, '=SUM(B14:B100)', ''],
            ['Membres Actifs', regular_members, '=B7', ''],
            ['Taux de Participation', f'{participation_rate}%', '=B15/(B16*B14)*100', 'Formule MDEVISP'],
        ]

        for row_idx, data in enumerate(cultes_data, 14):
            for col_idx, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = bold_font if col_idx == 1 else normal_font
                cell.alignment = center_align if col_idx > 1 else left_align
                cell.border = thin_border
                if col_idx == 1:
                    cell.fill = light_fill
                if row_idx == 17 and col_idx == 2:
                    cell.fill = highlight_fill

        # === SECTION 3: DETAILS PAR TYPE DE CULTE ===
        ws.merge_cells('A19:D19')
        ws['A19'].value = "3. DÉTAIL PAR TYPE DE CULTE"
        ws['A19'].font = header_font
        ws['A19'].fill = header_fill
        ws['A19'].alignment = center_align

        headers_types = ['Type de Culte', 'Présences', 'Pourcentage', 'Barre']
        for col, header in enumerate(headers_types, 1):
            cell = ws.cell(row=20, column=col, value=header)
            cell.font = bold_font
            cell.fill = PatternFill(start_color='93C5FD', end_color='93C5FD', fill_type='solid')
            cell.alignment = center_align
            cell.border = thin_border

        # Récupérer les cultes types
        culte_types = CulteType.objects.all()
        row_num = 21
        for ct in culte_types:
            count = attendances_qs.filter(culte_session__culte_type=ct).count()
            pct = round((count / total_presences * 100) if total_presences > 0 else 0, 1)
            
            ws.cell(row=row_num, column=1, value=ct.name).border = thin_border
            ws.cell(row=row_num, column=2, value=count).border = thin_border
            ws.cell(row=row_num, column=3, value=f'{pct}%').border = thin_border
            ws.cell(row=row_num, column=4, value=f'=B{row_num}/SUM($B$21:$B${row_num+10})*100').border = thin_border
            row_num += 1

        # === SECTION 4: FORMULES EXPLICATIVES ===
        ws.merge_cells('A35:D35')
        ws['A35'].value = "4. FORMULES DE CALCUL UTILISÉES"
        ws['A35'].font = header_font
        ws['A35'].fill = header_fill
        ws['A35'].alignment = center_align

        formules = [
            ['Taux de Participation', '=(Total Présences / (Membres Actifs × Nb Cultes)) × 100'],
            ['Pourcentage par Genre', '=(Effectif Genre / Total Membres) × 100'],
            ['Pourcentage par Type', '=(Présences Type / Total Présences) × 100'],
            ['Seuil d\'alerte', '< 40% = Participation insuffisante'],
        ]

        for row_idx, formule in enumerate(formules, 36):
            ws.cell(row=row_idx, column=1, value=formule[0]).font = bold_font
            ws.cell(row=row_idx, column=2, value=formule[1])
            ws.cell(row=row_idx, column=1).border = thin_border
            ws.cell(row=row_idx, column=2).border = thin_border

        # === PIED DE PAGE ===
        ws.merge_cells('A45:D45')
        ws['A45'].value = f"Document généré par D'avant Croissance - {timezone.now().strftime('%d/%m/%Y à %H:%M')}"
        ws['A45'].font = Font(name='Calibri', size=9, italic=True, color='666666')
        ws['A45'].alignment = center_align

        # Ajuster largeurs colonnes
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 25

        # Export
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="Statistiques_{user_church.name}_{date_from}_to_{date_to}.xlsx"'
        return response

    elif export_format == 'docx':
        # Export DOCX simple (HTML-like)
        content = f"""
        <html xmlns:o='urn:schemas-microsoft-com:office:office' xmlns:w='urn:schemas-microsoft-com:office:word'>
        <head><meta charset='utf-8'><title>Statistiques</title></head>
        <body>
            <h1>STATISTIQUES - {user_church.name}</h1>
            <p>Période: {date_from} au {date_to}</p>
            
            <h2>Membres</h2>
            <table border='1'>
                <tr><td>Total Membres</td><td>{total_members}</td></tr>
                <tr><td>Visiteurs</td><td>{visitors}</td></tr>
                <tr><td>Membres Actifs</td><td>{regular_members}</td></tr>
                <tr><td>Hommes</td><td>{hommes}</td></tr>
                <tr><td>Femmes</td><td>{femmes}</td></tr>
            </table>
            
            <h2>Cultes</h2>
            <table border='1'>
                <tr><td>Total Cultes</td><td>{total_cultes}</td></tr>
                <tr><td>Total Présences</td><td>{total_presences}</td></tr>
                <tr><td>Taux de Participation</td><td>{participation_rate}%</td></tr>
            </table>
        </body>
        </html>
        """
        
        response = HttpResponse(content_type='application/msword')
        response['Content-Disposition'] = f'attachment; filename="statistiques_{user_church.name}.doc"'
        response.write(content)
        return response
    
    else:  # pdf - on utilise l'impression navigateur
        messages.info(request, 'ℹ️ Pour exporter en PDF, utilisez la fonction d\'impression de votre navigateur (Ctrl+P)')
        return redirect('statistics')


@login_required(login_url='login')
def report_generate(request):
    """Générer un rapport administratif complet"""
    if not user_is_admin(request.user):
        messages.error(request, '⚠️ Accès réservé aux administrateurs uniquement.')
        return redirect('member_list')

    user_church = request.user.profile.church

    # Filtres
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')

    # Période par défaut (30 derniers jours)
    if not date_from:
        date_from = (timezone.now().date() - timedelta(days=30)).isoformat()
    if not date_to:
        date_to = timezone.now().date().isoformat()

    # Collecte des données
    sessions_qs = CulteSession.objects.filter(
        church=user_church,
        date__date__gte=date_from,
        date__date__lte=date_to
    ).select_related('culte_type').order_by('-date')
    
    attendances_qs = Attendance.objects.filter(
        member__church=user_church,
        culte_session__date__date__gte=date_from,
        culte_session__date__date__lte=date_to
    )

    # Statistiques générales
    total_members = Member.objects.filter(church=user_church).count()
    visitors = Member.objects.filter(church=user_church, status='VISITOR').count()
    regular_members = Member.objects.filter(church=user_church, status='REGULAR').count()
    total_cultes = sessions_qs.count()
    total_presences = attendances_qs.count()

    # Statistiques par genre
    hommes = Member.objects.filter(church=user_church, genre='H').count()
    femmes = Member.objects.filter(church=user_church, genre='F').count()
    enfants = Member.objects.filter(church=user_church, genre='E').count()

    # Nouveaux membres durant la période
    new_members = Member.objects.filter(
        church=user_church,
        date_joined__gte=date_from,
        date_joined__lte=date_to
    ).count()

    # Présences par type de culte
    culte_types = CulteType.objects.all()
    presences_by_type = []
    for ct in culte_types:
        count = attendances_qs.filter(culte_session__culte_type=ct).count()
        presences_by_type.append({
            'name': ct.name,
            'count': count,
            'percentage': round((count / total_presences * 100) if total_presences > 0 else 0, 1)
        })

    # Cultes avec plus de fréquentation (filtré par période)
    top_cultes = sessions_qs.annotate(
        attendance_count=Count('attendances', filter=Q(
            attendances__culte_session__date__date__gte=date_from,
            attendances__culte_session__date__date__lte=date_to
        ))
    ).order_by('-attendance_count')[:5]

    # Taux de participation (présences moyennes par culte / membres actifs)
    if regular_members > 0 and total_cultes > 0:
        participation_rate = round((total_presences / (regular_members * total_cultes)) * 100, 1)
    else:
        participation_rate = 0

    # Alertes d'absence non résolues
    active_alerts = AbsenceAlert.objects.filter(
        church=user_church,
        is_resolved=False
    ).count()

    context = {
        'user_church': user_church,
        'date_from': date_from,
        'date_to': date_to,
        'total_members': total_members,
        'visitors': visitors,
        'regular_members': regular_members,
        'total_cultes': total_cultes,
        'total_presences': total_presences,
        'hommes': hommes,
        'femmes': femmes,
        'enfants': enfants,
        'new_members': new_members,
        'presences_by_type': presences_by_type,
        'top_cultes': top_cultes,
        'participation_rate': participation_rate,
        'active_alerts': active_alerts,
        'report_date': timezone.now().date(),
    }

    return render(request, 'church/report_generate.html', context)


@login_required(login_url='login')
def report_export_pdf(request):
    """Exporter le rapport en PDF (via impression navigateur)"""
    if not user_is_admin(request.user):
        messages.error(request, '⚠️ Accès réservé aux administrateurs uniquement.')
        return redirect('member_list')

    # Rediriger vers la page de génération avec paramètre print
    return redirect('report_generate') + '?print=1'


@login_required(login_url='login')
def report_export_docx(request):
    """Exporter le rapport en DOCX"""
    if not user_is_admin(request.user):
        messages.error(request, '⚠️ Accès réservé aux administrateurs uniquement.')
        return redirect('member_list')

    user_church = request.user.profile.church
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')

    if not date_from:
        date_from = (timezone.now().date() - timedelta(days=30)).isoformat()
    if not date_to:
        date_to = timezone.now().date().isoformat()

    # Collecte des données (similaire à report_generate)
    total_members = Member.objects.filter(church=user_church).count()
    visitors = Member.objects.filter(church=user_church, status='VISITOR').count()
    regular_members = Member.objects.filter(church=user_church, status='REGULAR').count()
    hommes = Member.objects.filter(church=user_church, genre='H').count()
    femmes = Member.objects.filter(church=user_church, genre='F').count()
    enfants = Member.objects.filter(church=user_church, genre='E').count()
    
    total_cultes = CulteSession.objects.filter(
        church=user_church,
        date__date__gte=date_from,
        date__date__lte=date_to
    ).count()
    total_presences = Attendance.objects.filter(
        member__church=user_church,
        culte_session__date__date__gte=date_from,
        culte_session__date__date__lte=date_to
    ).count()
    
    # Calcul du taux de participation
    participation_rate = round((total_presences / (regular_members * total_cultes) * 100) if regular_members > 0 and total_cultes > 0 else 0, 1)
    
    # Nouveaux membres
    new_members = Member.objects.filter(
        church=user_church,
        date_joined__gte=date_from,
        date_joined__lte=date_to
    ).count()
    
    # Alertes
    active_alerts = AbsenceAlert.objects.filter(
        church=user_church,
        is_resolved=False
    ).count()

    # Top cultes
    sessions_qs = CulteSession.objects.filter(
        church=user_church,
        date__date__gte=date_from,
        date__date__lte=date_to
    ).select_related('culte_type').order_by('-date')
    top_cultes = sessions_qs.annotate(
        attendance_count=Count('attendances', filter=Q(
            attendances__culte_session__date__date__gte=date_from,
            attendances__culte_session__date__date__lte=date_to
        ))
    ).order_by('-attendance_count')[:5]

    # Pourcentages
    pct_hommes = round((hommes / total_members * 100) if total_members > 0 else 0, 1)
    pct_femmes = round((femmes / total_members * 100) if total_members > 0 else 0, 1)
    pct_enfants = round((enfants / total_members * 100) if total_members > 0 else 0, 1)
    avg_presences = round(total_presences / total_cultes, 1) if total_cultes > 0 else 0

    # Format DOCX - HTML simple compatible Word avec UTF-8
    report_date = timezone.now()
    
    content = f"""<!DOCTYPE html>
<html>
<head>
    <meta http-equiv="Content-Type" content="text/html; charset=UTF-8">
    <meta charset="UTF-8">
    <title>Rapport - {user_church.name}</title>
    <style>
        @page {{ margin: 2cm; }}
        body {{ font-family: 'Calibri', Arial, sans-serif; font-size: 12pt; color: #000; line-height: 1.6; }}
        .header {{ background: #1e40af; color: white; padding: 30px; text-align: center; margin-bottom: 30px; }}
        .header h1 {{ color: white; font-size: 24pt; margin: 0 0 15px 0; border-bottom: 2px solid white; padding-bottom: 15px; }}
        .header h2 {{ color: white; font-size: 18pt; margin: 10px 0; background: none; border: none; padding: 0; }}
        .header p {{ margin: 8px 0; font-size: 11pt; }}
        h1 {{ color: #1e3a8a; font-size: 20pt; font-weight: bold; border-bottom: 3px solid #3b82f6; padding-bottom: 10px; margin-top: 30px; }}
        h2 {{ color: #1e40af; font-size: 16pt; font-weight: bold; background: #eff6ff; padding: 10px; border-left: 4px solid #3b82f6; margin-top: 25px; }}
        h3 {{ color: #333; font-size: 14pt; font-weight: bold; margin-top: 20px; }}
        table {{ border-collapse: collapse; width: 100%; margin: 15px 0; }}
        th {{ border: 2px solid #3b82f6; padding: 10px; background: #3b82f6; color: white; font-weight: bold; text-align: center; }}
        td {{ border: 1px solid #999; padding: 8px; text-align: center; }}
        tr:nth-child(even) {{ background: #f5f5f5; }}
        .highlight {{ background: #fef3c7 !important; font-weight: bold; }}
        .success {{ color: #059669; font-weight: bold; }}
        .warning {{ color: #d97706; font-weight: bold; }}
        .stat-table {{ margin: 20px 0; }}
        .stat-table td {{ border: 2px solid #3b82f6; padding: 15px; text-align: center; width: 25%; }}
        .stat-number {{ font-size: 32pt; font-weight: bold; color: #3b82f6; }}
        .stat-label {{ font-size: 10pt; color: #666; text-transform: uppercase; }}
        .info-box {{ background: #eff6ff; padding: 15px; border-left: 4px solid #3b82f6; margin: 20px 0; }}
        .footer {{ margin-top: 40px; text-align: center; font-size: 10pt; color: #666; border-top: 2px solid #ddd; padding-top: 15px; }}
        .signature {{ margin-top: 50px; width: 100%; }}
        .signature td {{ border: none !important; border-top: 3px solid #3b82f6 !important; padding: 20px; width: 50%; background: none !important; }}
        .page-break {{ page-break-before: always; }}
    </style>
</head>
<body>
    <!-- En-tete -->
    <div class='header'>
        <h1>RAPPORT ADMINISTRATIF</h1>
        <h2>{user_church.name}</h2>
        <p><strong>Localisation :</strong> {user_church.location.name if user_church.location else 'N/A'}</p>
        <p><strong>Periode :</strong> du {date_from} au {date_to}</p>
        <p><em>Genere le {report_date.strftime('%d/%m/%Y a %H:%M')}</em></p>
    </div>

    <!-- Section 1 -->
    <h1>1. STATISTIQUES GENERALES</h1>
    
    <table class='stat-table'>
        <tr>
            <td>
                <div class='stat-number'>{total_members}</div>
                <div class='stat-label'>Total Membres</div>
            </td>
            <td>
                <div class='stat-number' style='color:#059669;'>{regular_members}</div>
                <div class='stat-label'>Membres Actifs</div>
            </td>
            <td>
                <div class='stat-number' style='color:#d97706;'>{visitors}</div>
                <div class='stat-label'>Visiteurs</div>
            </td>
            <td>
                <div class='stat-number' style='color:#7c3aed;'>{total_cultes}</div>
                <div class='stat-label'>Cultes</div>
            </td>
        </tr>
    </table>

    <h3>Repartition par Genre</h3>
    <table>
        <tr>
            <th style='width:40%;'>Genre</th>
            <th style='width:30%;'>Effectif</th>
            <th style='width:30%;'>Pourcentage</th>
        </tr>
        <tr>
            <td style='text-align:left; padding-left:15px;'>Hommes</td>
            <td><b>{hommes}</b></td>
            <td>{pct_hommes}%</td>
        </tr>
        <tr>
            <td style='text-align:left; padding-left:15px;'>Femmes</td>
            <td><b>{femmes}</b></td>
            <td>{pct_femmes}%</td>
        </tr>
        <tr>
            <td style='text-align:left; padding-left:15px;'>Enfants</td>
            <td><b>{enfants}</b></td>
            <td>{pct_enfants}%</td>
        </tr>
        <tr class='highlight'>
            <td style='text-align:left; padding-left:15px;'><b>TOTAL</b></td>
            <td><b>{total_members}</b></td>
            <td><b>100%</b></td>
        </tr>
    </table>

    <div class='info-box'>
        <p><strong>Informations :</strong></p>
        <p>Nouveaux membres : <span class='success'>{new_members}</span></p>
        <p>Alertes d'absence : <span class='warning'>{active_alerts}</span></p>
    </div>

    <!-- Section 2 -->
    <h1>2. ANALYSE DES PRESENCES</h1>
    
    <table>
        <tr>
            <th style='width:65%; text-align:left; padding-left:15px;'>Indicateur</th>
            <th style='width:35%;'>Valeur</th>
        </tr>
        <tr>
            <td style='text-align:left; padding-left:15px;'>Total des cultes</td>
            <td><b>{total_cultes}</b></td>
        </tr>
        <tr>
            <td style='text-align:left; padding-left:15px;'>Total des presences</td>
            <td><b>{total_presences}</b></td>
        </tr>
        <tr class='highlight'>
            <td style='text-align:left; padding-left:15px;'><strong>Taux de participation</strong></td>
            <td><span class='success'><b>{participation_rate}%</b></span></td>
        </tr>
        <tr>
            <td style='text-align:left; padding-left:15px;'>Moyenne par culte</td>
            <td><b>{avg_presences}</b></td>
        </tr>
    </table>

    <!-- Section 3 -->
    <div class='page-break'></div>
    <h1>3. CULTES LES PLUS FREQUENTES</h1>
    
    <table>
        <tr>
            <th style='width:12%;'>Rang</th>
            <th style='width:20%;'>Date</th>
            <th style='width:22%;'>Type</th>
            <th style='width:31%;'>Theme</th>
            <th style='width:15%;'>Presences</th>
        </tr>
"""

    # Top cultes
    medals = ['1er', '2eme', '3eme']
    for i, culte in enumerate(top_cultes):
        rang = medals[i] if i < 3 else f'{i+1}eme'
        highlight = " class='highlight'" if i < 3 else ""
        content += f"""        <tr{highlight}>
            <td style='font-size:12pt;'>{rang}</td>
            <td>{culte.date.strftime('%d/%m/%Y')}</td>
            <td>{culte.culte_type.name}</td>
            <td style='text-align:left; padding-left:10px;'>{culte.theme or '—'}</td>
            <td><b>{culte.attendance_count}</b></td>
        </tr>
"""

    if not top_cultes:
        content += """        <tr>
            <td colspan='5' style='text-align:center;padding:30px;color:#999;'><em>Aucun culte enregistre</em></td>
        </tr>
"""

    content += f"""
    </table>

    <!-- Section 4 -->
    <h1>4. CONCLUSION</h1>
    
    <div class='info-box'>
        <p><strong>Analyse de la periode :</strong></p>
        <p>Ce rapport presente les statistiques de l'eglise <strong>{user_church.name}</strong> pour la periode du <strong>{date_from}</strong> au <strong>{date_to}</strong>.</p>
        
        <p><strong>Points cles :</strong></p>
        <ul style='line-height:1.8;'>
            <li>L'eglise compte <strong>{total_members} membres</strong> dont <strong>{regular_members} membres actifs</strong></li>
            <li>Le taux de participation moyen est de <span class='success'><strong>{participation_rate}%</strong></span></li>
            <li><span class='success'><strong>+{new_members} nouveau(x) membre(s)</strong></span> ont rejoint l'eglise</li>
            <li><span class='warning'><strong>{active_alerts} membre(s)</strong></span> necessitent un suivi pastoral</li>
        </ul>
        
        <p style='margin-top:15px; padding:10px; background:#f0fdf4; border-left:3px solid #059669;'>
            <em>Pour toute question, veuillez contacter l'administration de l'eglise.</em>
        </p>
    </div>

    <!-- Signatures -->
    <div class='signature'>
        <table>
            <tr>
                <td>
                    <p style='margin:0; font-weight:bold; font-size:12pt;'>L'Administrateur</p>
                    <p style='margin:10px 0 0 0; color:#666; font-size:10pt;'>Date et signature</p>
                </td>
                <td>
                    <p style='margin:0; font-weight:bold; font-size:12pt;'>Le Pasteur Responsable</p>
                    <p style='margin:10px 0 0 0; color:#666; font-size:10pt;'>Date et signature</p>
                </td>
            </tr>
        </table>
    </div>

    <!-- Pied de page -->
    <div class='footer'>
        <p><strong>Document genere par D'avant Croissance</strong></p>
        <p>Systeme de Gestion Intelligente d'Eglise</p>
        <p style='margin-top:10px; font-size:9pt;'>C {report_date.year} - Tous droits reserves | Document confidentiel</p>
        <p style='font-size:9pt;'>Genere le {report_date.strftime('%d/%m/%Y a %H:%M:%S')}</p>
    </div>
</body>
</html>
"""

    response = HttpResponse(content_type='application/vnd.ms-word')
    response['Content-Disposition'] = f'attachment; filename="Rapport_{user_church.name}_{date_from}_to_{date_to}.doc"'
    response['Content-Type'] = 'application/vnd.ms-word; charset=UTF-8'
    response.write(content)
    return response


# ==========================================================
# RAPPORTS MDEVISP - Système de Suivi des Cultes
# ==========================================================

@login_required(login_url='login')
def mdevisp_report(request):
    """Rapport mensuel détaillé MDEVISP avec calculs officiels"""
    if not user_is_admin(request.user):
        messages.error(request, '⚠️ Accès réservé aux administrateurs uniquement.')
        return redirect('member_list')

    user_church = request.user.profile.church
    
    # Récupérer l'année et le mois (défaut : mois actuel)
    year = int(request.GET.get('year', timezone.now().year))
    month = int(request.GET.get('month', timezone.now().month))
    
    # Calcul des périodes
    current_month_start = date(year, month, 1)
    if month == 12:
        current_month_end = date(year + 1, 1, 1) - timedelta(days=1)
        previous_month_start = date(year, month - 1, 1)
        previous_month_end = date(year, month, 1) - timedelta(days=1)
    else:
        current_month_end = date(year, month + 1, 1) - timedelta(days=1)
        previous_month_start = date(year, month - 1, 1) if month > 1 else date(year - 1, 12, 1)
        previous_month_end = date(year, month, 1) - timedelta(days=1)
    
    # Récupérer les types de cultes MDEVISP
    culte_matinal = CulteType.objects.filter(name__icontains='matinal').first()
    culte_vesperal = CulteType.objects.filter(name__icontains='vesp').first()
    culte_dominical = CulteType.objects.filter(name__icontains='domin').first()
    
    # Si les types n'existent pas, on utilise tous les cultes
    culte_types_filter = []
    if culte_matinal:
        culte_types_filter.append(culte_matinal.id)
    if culte_vesperal:
        culte_types_filter.append(culte_vesperal.id)
    if culte_dominical:
        culte_types_filter.append(culte_dominical.id)
    
    # Si aucun type MDEVISP trouvé, on prend tous les cultes
    if not culte_types_filter:
        all_types = CulteType.objects.all()
        culte_types_filter = [ct.id for ct in all_types]
    
    # === STATISTIQUES MOIS ACTUEL ===
    sessions_actuel = CulteSession.objects.filter(
        church=user_church,
        date__date__gte=current_month_start,
        date__date__lte=current_month_end,
        culte_type_id__in=culte_types_filter
    )
    
    # Comptage par type de culte
    cultes_matinaux = sessions_actuel.filter(culte_type=culte_matinal).count() if culte_matinal else 0
    cultes_vesperaux = sessions_actuel.filter(culte_type=culte_vesperal).count() if culte_vesperal else 0
    cultes_dominicaux = sessions_actuel.filter(culte_type=culte_dominical).count() if culte_dominical else 0
    
    # Total X = N + M + T
    total_cultes_actuel = cultes_matinaux + cultes_vesperaux + cultes_dominicaux
    
    # Présences totales du mois
    presences_actuel = Attendance.objects.filter(
        member__church=user_church,
        culte_session__in=sessions_actuel
    ).count()
    
    # Effectif total Y (membres actifs)
    effectif_total = Member.objects.filter(church=user_church, status='REGULAR').count()
    
    # Calcul du pourcentage mensuel : % = (X / Y) × 100
    # Mais on utilise plutôt : % = (Présences / (Effectif × Nb cultes)) × 100
    if effectif_total > 0 and total_cultes_actuel > 0:
        pourcentage_mensuel = round((presences_actuel / (effectif_total * total_cultes_actuel)) * 100, 1)
    else:
        pourcentage_mensuel = 0
    
    # === STATISTIQUES MOIS PRÉCÉDENT ===
    sessions_previous = CulteSession.objects.filter(
        church=user_church,
        date__date__gte=previous_month_start,
        date__date__lte=previous_month_end,
        culte_type_id__in=culte_types_filter
    )
    
    cultes_matinaux_prev = sessions_previous.filter(culte_type=culte_matinal).count() if culte_matinal else 0
    cultes_vesperaux_prev = sessions_previous.filter(culte_type=culte_vesperal).count() if culte_vesperal else 0
    cultes_dominicaux_prev = sessions_previous.filter(culte_type=culte_dominical).count() if culte_dominical else 0
    total_cultes_previous = cultes_matinaux_prev + cultes_vesperaux_prev + cultes_dominicaux_prev
    
    presences_previous = Attendance.objects.filter(
        member__church=user_church,
        culte_session__in=sessions_previous
    ).count()
    
    if effectif_total > 0 and total_cultes_previous > 0:
        pourcentage_previous = round((presences_previous / (effectif_total * total_cultes_previous)) * 100, 1)
    else:
        pourcentage_previous = 0
    
    # Calcul de la tendance
    evolution = pourcentage_mensuel - pourcentage_previous
    tendance = "stable"
    if evolution > 2:
        tendance = "hausse"
    elif evolution < -2:
        tendance = "baisse"
    
    # === STATISTIQUES PAR GENRE ===
    hommes = Member.objects.filter(church=user_church, genre='H').count()
    femmes = Member.objects.filter(church=user_church, genre='F').count()
    enfants = Member.objects.filter(church=user_church, genre='E').count()
    
    # === STATISTIQUES PAR STATUT ===
    visiteurs = Member.objects.filter(church=user_church, status='VISITOR').count()
    membres_actifs = Member.objects.filter(church=user_church, status='REGULAR').count()
    
    # === ALERTES D'ABSENCE ===
    config = GrowthConfig.get_config()
    seuil_absence = timezone.now() - timedelta(days=config.absence_limit_days)
    
    alertes_absence = AbsenceAlert.objects.filter(
        church=user_church,
        is_resolved=False
    ).count()
    
    # === SEUIL D'ALERTE ===
    seuil_alerte = 40  # 40% selon cahier des charges
    alerte_seuil = pourcentage_mensuel < seuil_alerte
    
    # === DONNÉES POUR GRAPHIQUES ===
    # Évolution sur les 6 derniers mois
    monthly_data = []
    for i in range(5, -1, -1):
        ref_date = timezone.now() - relativedelta(months=i)
        month_start = date(ref_date.year, ref_date.month, 1)
        month_end = date(ref_date.year, ref_date.month + 1, 1) - timedelta(days=1) if ref_date.month < 12 else date(ref_date.year, 12, 31)
        
        sessions = CulteSession.objects.filter(
            church=user_church,
            date__date__gte=month_start,
            date__date__lte=month_end
        )
        presences = Attendance.objects.filter(member__church=user_church, culte_session__in=sessions).count()
        nb_cultes = sessions.count()
        
        if effectif_total > 0 and nb_cultes > 0:
            pct = round((presences / (effectif_total * nb_cultes)) * 100, 1)
        else:
            pct = 0
        
        monthly_data.append({
            'month': ref_date.strftime('%b %Y'),
            'percentage': pct,
            'presences': presences,
            'cultes': nb_cultes
        })
    
    context = {
        'user_church': user_church,
        'year': year,
        'month': month,
        'month_name': current_month_start.strftime('%B %Y'),
        
        # Données mois actuel
        'cultes_matinaux': cultes_matinaux,
        'cultes_vesperaux': cultes_vesperaux,
        'cultes_dominicaux': cultes_dominicaux,
        'total_cultes_actuel': total_cultes_actuel,
        'presences_actuel': presences_actuel,
        'effectif_total': effectif_total,
        'pourcentage_mensuel': pourcentage_mensuel,
        
        # Données mois précédent
        'pourcentage_previous': pourcentage_previous,
        'evolution': evolution,
        'tendance': tendance,
        
        # Statistiques détaillées
        'hommes': hommes,
        'femmes': femmes,
        'enfants': enfants,
        'visiteurs': visiteurs,
        'membres_actifs': membres_actifs,
        
        # Alertes
        'alertes_absence': alertes_absence,
        'seuil_alerte': seuil_alerte,
        'alerte_seuil': alerte_seuil,
        
        # Graphiques
        'monthly_data': monthly_data,
        
        # Navigation
        'previous_month': month - 1 if month > 1 else 12,
        'previous_year': year if month > 1 else year - 1,
        'next_month': month + 1 if month < 12 else 1,
        'next_year': year if month < 12 else year + 1,
    }
    
    return render(request, 'church/mdevisp_report.html', context)


@login_required(login_url='login')
def mdevisp_annual_report(request):
    """Rapport annuel consolidé MDEVISP"""
    if not user_is_admin(request.user):
        messages.error(request, '⚠️ Accès réservé aux administrateurs uniquement.')
        return redirect('member_list')

    user_church = request.user.profile.church
    year = int(request.GET.get('year', timezone.now().year))
    
    # Données mensuelles pour l'année
    annual_data = []
    total_presences_year = 0
    total_cultes_year = 0
    
    for month in range(1, 13):
        month_start = date(year, month, 1)
        month_end = date(year, month + 1, 1) - timedelta(days=1) if month < 12 else date(year, 12, 31)
        
        sessions = CulteSession.objects.filter(
            church=user_church,
            date__date__gte=month_start,
            date__date__lte=month_end
        )
        presences = Attendance.objects.filter(member__church=user_church, culte_session__in=sessions).count()
        nb_cultes = sessions.count()
        
        effectif = Member.objects.filter(church=user_church, status='REGULAR').count()
        
        if effectif > 0 and nb_cultes > 0:
            pct = round((presences / (effectif * nb_cultes)) * 100, 1)
        else:
            pct = 0
        
        annual_data.append({
            'month': month_start.strftime('%B'),
            'percentage': pct,
            'presences': presences,
            'cultes': nb_cultes
        })
        
        total_presences_year += presences
        total_cultes_year += nb_cultes
    
    # Moyenne annuelle
    valid_months = [d['percentage'] for d in annual_data if d['percentage'] > 0]
    moyenne_annuelle = round(sum(valid_months) / len(valid_months), 1) if valid_months else 0
    
    # Effectif actuel
    effectif_total = Member.objects.filter(church=user_church, status='REGULAR').count()
    
    # Taux de participation annuel
    if effectif_total > 0 and total_cultes_year > 0:
        participation_annuel = round((total_presences_year / (effectif_total * total_cultes_year)) * 100, 1)
    else:
        participation_annuel = 0
    
    context = {
        'user_church': user_church,
        'year': year,
        'annual_data': annual_data,
        'moyenne_annuelle': moyenne_annuelle,
        'total_presences_year': total_presences_year,
        'total_cultes_year': total_cultes_year,
        'effectif_total': effectif_total,
        'participation_annuel': participation_annuel,
    }
    
    return render(request, 'church/mdevisp_annual.html', context)


@login_required(login_url='login')
def mdevisp_export_docx(request):
    """Exporter le rapport MDEVISP en DOCX"""
    if not user_is_admin(request.user):
        messages.error(request, '⚠️ Accès réservé aux administrateurs uniquement.')
        return redirect('member_list')

    user_church = request.user.profile.church
    year = int(request.GET.get('year', timezone.now().year))
    month = int(request.GET.get('month', timezone.now().month))
    
    # Calcul des périodes (même logique que mdevisp_report)
    current_month_start = date(year, month, 1)
    if month == 12:
        current_month_end = date(year + 1, 1, 1) - timedelta(days=1)
    else:
        current_month_end = date(year, month + 1, 1) - timedelta(days=1)
    
    # Récupérer les cultes
    culte_matinal = CulteType.objects.filter(name__icontains='matinal').first()
    culte_vesperal = CulteType.objects.filter(name__icontains='vesp').first()
    culte_dominical = CulteType.objects.filter(name__icontains='domin').first()
    
    sessions = CulteSession.objects.filter(
        church=user_church,
        date__date__gte=current_month_start,
        date__date__lte=current_month_end
    ).select_related('culte_type').order_by('date')
    
    # Statistiques
    cultes_matinaux = sessions.filter(culte_type=culte_matinal).count() if culte_matinal else 0
    cultes_vesperaux = sessions.filter(culte_type=culte_vesperal).count() if culte_vesperal else 0
    cultes_dominicaux = sessions.filter(culte_type=culte_dominical).count() if culte_dominical else 0
    total_cultes = cultes_matinaux + cultes_vesperaux + cultes_dominicaux
    
    presences = Attendance.objects.filter(member__church=user_church, culte_session__in=sessions).count()
    effectif = Member.objects.filter(church=user_church, status='REGULAR').count()
    
    if effectif > 0 and total_cultes > 0:
        pourcentage = round((presences / (effectif * total_cultes)) * 100, 1)
    else:
        pourcentage = 0
    
    hommes = Member.objects.filter(church=user_church, genre='H').count()
    femmes = Member.objects.filter(church=user_church, genre='F').count()
    enfants = Member.objects.filter(church=user_church, genre='E').count()
    
    alertes = AbsenceAlert.objects.filter(church=user_church, is_resolved=False).count()
    
    report_date = timezone.now()
    month_name = current_month_start.strftime('%B %Y')
    
    content = f"""<!DOCTYPE html>
<html>
<head>
    <meta http-equiv="Content-Type" content="text/html; charset=UTF-8">
    <meta charset="UTF-8">
    <title>Rapport MDEVISP - {month_name}</title>
    <style>
        @page {{ margin: 2cm; }}
        body {{ font-family: 'Calibri', Arial, sans-serif; font-size: 11pt; color: #000; }}
        .header {{ background: #1e40af; color: white; padding: 25px; text-align: center; margin-bottom: 25px; }}
        .header h1 {{ color: white; font-size: 20pt; margin: 0 0 10px 0; }}
        .header h2 {{ color: white; font-size: 16pt; margin: 5px 0; }}
        h1 {{ color: #1e3a8a; font-size: 18pt; font-weight: bold; border-bottom: 3px solid #3b82f6; padding-bottom: 8px; margin-top: 25px; }}
        h2 {{ color: #1e40af; font-size: 14pt; font-weight: bold; background: #eff6ff; padding: 8px; margin-top: 20px; }}
        table {{ border-collapse: collapse; width: 100%; margin: 15px 0; }}
        th {{ border: 2px solid #3b82f6; padding: 8px; background: #3b82f6; color: white; font-weight: bold; text-align: center; }}
        td {{ border: 1px solid #999; padding: 6px; text-align: center; }}
        tr:nth-child(even) {{ background: #f5f5f5; }}
        .highlight {{ background: #fef3c7 !important; font-weight: bold; }}
        .success {{ color: #059669; font-weight: bold; }}
        .warning {{ color: #d97706; font-weight: bold; }}
        .info-box {{ background: #eff6ff; padding: 12px; border-left: 4px solid #3b82f6; margin: 15px 0; }}
        .stat-grid {{ display: table; width: 100%; margin: 15px 0; }}
        .stat-row {{ display: table-row; }}
        .stat-cell {{ display: table-cell; border: 2px solid #3b82f6; padding: 12px; text-align: center; width: 25%; }}
        .stat-number {{ font-size: 24pt; font-weight: bold; color: #3b82f6; }}
        .stat-label {{ font-size: 9pt; color: #666; text-transform: uppercase; }}
        .footer {{ margin-top: 30px; text-align: center; font-size: 9pt; color: #666; border-top: 2px solid #ddd; padding-top: 10px; }}
    </style>
</head>
<body>
    <div class='header'>
        <h1>EGLISE EVANGELIQUE DU CONGO</h1>
        <h2>Rapport MDEVISP - {month_name}</h2>
        <p>Paroisse : {user_church.name} - {user_church.location.name if user_church.location else 'N/A'}</p>
    </div>

    <h1>1. STATISTIQUES MENSUELLES</h1>
    
    <table class='stat-grid'>
        <tr class='stat-row'>
            <td class='stat-cell'>
                <div class='stat-number'>{cultes_matinaux}</div>
                <div class='stat-label'>Cultes Matinaux (N)</div>
            </td>
            <td class='stat-cell'>
                <div class='stat-number'>{cultes_vesperaux}</div>
                <div class='stat-label'>Cultes Vesperaux (M)</div>
            </td>
            <td class='stat-cell'>
                <div class='stat-number'>{cultes_dominicaux}</div>
                <div class='stat-label'>Cultes Dominicaux (T)</div>
            </td>
            <td class='stat-cell'>
                <div class='stat-number' style='color:#7c3aed;'>{total_cultes}</div>
                <div class='stat-label'>Total X = N+M+T</div>
            </td>
        </tr>
    </table>

    <table>
        <tr>
            <th style='width:50%; text-align:left; padding-left:10px;'>Indicateur</th>
            <th style='width:50%;'>Valeur</th>
        </tr>
        <tr>
            <td style='text-align:left; padding-left:10px;'>Effectif total (Y)</td>
            <td><b>{effectif}</b></td>
        </tr>
        <tr>
            <td style='text-align:left; padding-left:10px;'>Total presences</td>
            <td><b>{presences}</b></td>
        </tr>
        <tr class='highlight'>
            <td style='text-align:left; padding-left:10px;'><strong>Pourcentage mensuel</strong></td>
            <td><span class='success'><b>{pourcentage}%</b></span></td>
        </tr>
        <tr>
            <td style='text-align:left; padding-left:10px;'>Formule</td>
            <td><i>% = (X / Y) x 100</i></td>
        </tr>
    </table>

    <h1>2. STATISTIQUES PAR GENRE</h1>
    <table>
        <tr>
            <th style='width:33%;'>Hommes</th>
            <th style='width:33%;'>Femmes</th>
            <th style='width:33%;'>Enfants</th>
        </tr>
        <tr>
            <td><b>{hommes}</b></td>
            <td><b>{femmes}</b></td>
            <td><b>{enfants}</b></td>
        </tr>
    </table>

    <h1>3. ALERTES</h1>
    <div class='info-box'>
        <p><strong>Membres absents depuis plus de 30 jours :</strong> <span class='warning'>{alertes}</span></p>
        {(f"<p><strong>Alerte : Pourcentage en dessous du seuil de 40% !</strong></p>" if pourcentage < 40 else "<p><strong>Statut :</strong> <span class='success'>Pourcentage acceptable</span></p>")}
    </div>

    <h1>4. DETAILS DES CULTES</h1>
    <table>
        <tr>
            <th style='width:20%;'>Date</th>
            <th style='width:25%;'>Type</th>
            <th style='width:35%;'>Theme</th>
            <th style='width:20%;'>Presences</th>
        </tr>
"""

    for session in sessions:
        attendance_count = Attendance.objects.filter(culte_session=session).count()
        content += f"""        <tr>
            <td>{session.date.strftime('%d/%m/%Y %H:%M')}</td>
            <td>{session.culte_type.name}</td>
            <td style='text-align:left; padding-left:5px;'>{session.theme or '—'}</td>
            <td><b>{attendance_count}</b></td>
        </tr>
"""

    content += f"""
    </table>

    <div class='footer'>
        <p><strong>Document genere par D'avant Croissance - Systeme MDEVISP</strong></p>
        <p>Eglise Evangelique du Congo | {report_date.strftime('%d/%m/%Y a %H:%M')}</p>
    </div>
</body>
</html>
"""

    response = HttpResponse(content_type='application/vnd.ms-word')
    response['Content-Disposition'] = f'attachment; filename="MDEVISP_{user_church.name}_{month_name}.doc"'
    response['Content-Type'] = 'application/vnd.ms-word; charset=UTF-8'
    response.write(content)
    return response
